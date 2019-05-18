from keras.layers import Input, Dense
from keras.models import Model
import matplotlib.pyplot as plt
import numpy as np
np.random.seed(1337)

from keras.datasets import mnist
import pandas as pd
import csv
import math
import random
import xlrd
import matplotlib.pyplot as plt
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from numpy import *
from sklearn.model_selection import train_test_split
from sklearn.metrics import roc_curve
from sklearn.metrics import roc_auc_score
import numpy as np


# 定义函数
def ReadMyCsv(SaveList, fileName):
    csv_reader = csv.reader(open(fileName))
    for row in csv_reader:  # 把每个rna疾病对加入OriginalData，注意表头
        SaveList.append(row)
    return

def storFile(data, fileName):
    with open(fileName, "w", newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(data)
    return

# 读取源文件
OriginalData = []
ReadMyCsv(OriginalData, "drug-disease-whole.csv")
print(len(OriginalData))

# mesh = np.loadtxt("MeshID_original-whole.txt", dtype=str, delimiter=";")  # 读取源文件
# print('mesh长度', len(mesh))


# 预处理
# 小写OriginalData
counter = 0
while counter < len(OriginalData):
    OriginalData[counter][0] = OriginalData[counter][0].lower()
    OriginalData[counter][1] = OriginalData[counter][1].lower()
    counter = counter + 1
print('小写OriginalData')




LncDisease = []
counter = 0
while counter < len(OriginalData):
    Pair = []
    Pair.append(OriginalData[counter][0])
    Pair.append(OriginalData[counter][1])
    LncDisease.append(Pair)
    counter = counter + 1
print('LncDisease的长度', len(LncDisease))
print('OriginalData的长度', len(OriginalData))



# 构建AllDisease
AllDisease = []
counter1 = 0
while counter1 < len(OriginalData): #顺序遍历原始数据，构建AllDisease
    counter2 = 0
    flag = 0
    while counter2 < len(AllDisease):  #遍历AllDisease
        if OriginalData[counter1][1] != AllDisease[counter2]:#有新疾病
            counter2 = counter2 + 1
        elif OriginalData[counter1][1] == AllDisease[counter2]:#没有新疾病，用两个if第二个if会越界
            flag = 1
            counter2 = counter2 + 1
    if flag == 0:
        AllDisease.append(OriginalData[counter1][1])
    counter1 = counter1 + 1
print('len(AllDisease)', len(AllDisease))
# storFile(AllDisease, 'AllDisease.csv')
# AllDisease1 = AllDisease    # 为生成负样本


# 构建AllDRUG
AllDRUG = []
counter1 = 0
while counter1 < len(OriginalData): #顺序遍历原始数据，构建AllDisease
    counter2 = 0
    flag = 0
    while counter2 < len(AllDRUG):  #遍历AllDisease
        if OriginalData[counter1][0] != AllDRUG[counter2]:#有新疾病
            counter2 = counter2 + 1
        elif OriginalData[counter1][0] == AllDRUG[counter2]:#没有新疾病，用两个if第二个if会越界
            flag = 1
            break
    if flag == 0:
        AllDRUG.append(OriginalData[counter1][0])
    counter1 = counter1 + 1
print('len(AllDRUG)', len(AllDRUG))
storFile(AllDRUG, 'AllDRUG.csv')



# 由drug-disease生成对应关系矩阵，有关系1，没关系0，行为疾病AllDisease，列为 AllDRUG
# 生成全0矩阵
DiseaseAndDrugBinary = []
counter = 0
while counter < len(AllDisease):
    row = []
    counter1 = 0
    while counter1 < len(AllDRUG):
        row.append(0)
        counter1 = counter1 + 1
    DiseaseAndDrugBinary.append(row)
    counter = counter + 1


print('len(LncDisease)', len(LncDisease))
counter = 0
while counter < len(LncDisease):
    DN = LncDisease[counter][1]
    RN = LncDisease[counter][0]
    counter1 = 0
    while counter1 < len(AllDisease):
        if AllDisease[counter1] == DN:
            counter2 = 0
            while counter2 < len(AllDRUG):
                if AllDRUG[counter2] == RN:
                    DiseaseAndDrugBinary[counter1][counter2] = 1
                    break
                counter2 = counter2 + 1
            break
        counter1 = counter1 + 1
    counter = counter + 1
print('len(DiseaseAndDrugBinary)', len(DiseaseAndDrugBinary))
# storFile(DiseaseAndDrugBinary, 'DiseaseAndDrugBinary')


# 构建疾病的DAGs
# 构建dags的根节点
# DAGs = []
# counter1 = 0
# while counter1 < len(AllDisease):
#     group = []
#     group.extend(DiseaseAndMeshID[counter1])
#     group.append(0)
#     group1 = []
#     group1.append(group)
#     DAGs.append(group1)
#     counter1 = counter1 + 1
# print('len(DAGs)的叶子', len(DAGs))

# 生成AllDisease的完整DGAs[[[RootDisease,[ID,ID],layer],[FatherDisease,[ID,ID],layer]...],...]
# counter = 0
# while counter < len(DAGs):
#     if DAGs[counter][0][1] == 0:
#         counter = counter + 1
#         continue
#     counter1 = 0
#     while counter1 < len(DAGs[counter]):  #################
#         counter2 = 0
#         while counter2 < len(DAGs[counter][counter1][1]):  ###################只对一个节点扩展只能生成的二层信息
#             layer = DAGs[counter][counter1][2]  #######################
#             # if len(DAGs[0][counter1][1][counter2]) <= 3:
#             #     break
#             if len(DAGs[counter][counter1][1][counter2]) > 3:  ####################
#                 NID = DAGs[counter][counter1][1][counter2]  #####################
#                 L = len(NID)
#                 NID = NID[0:L - 4]  # 把id减3
#                 counter3 = 0
#                 flag = 1  # 默认不在
#                 while counter3 < len(mesh):  # 判断nid是否在mesh中，如果在求出疾病名，如果不在，跳出循还
#                     if NID == mesh[counter3][1]:
#                         flag = 0  # 由counter3找对应的疾病名
#                         num = counter3
#                         DiseaseName = mesh[counter3][0]
#                         break
#                     counter3 = counter3 + 1
#
#                 flag2 = 0  # 默认在dags不存在
#                 counter5 = 0
#                 while counter5 < len(DAGs[counter]):  # 找到对应疾病的名字后查找dags看是否已经出现，出现了就不加了
#                     if DAGs[counter][counter5][0] == DiseaseName:  #########################
#                         flag2 = 1  # dags中出现了
#                         break
#                     counter5 = counter5 + 1
#
#                 if flag == 0:
#                     if flag2 == 0:
#                         counter6 = 0    # 遍历mesh，寻找disease对应的id
#                         IDGroup = []
#                         while counter6 < len(mesh):
#                             if DiseaseName == mesh[counter6][0]:
#                                 IDGroup.append(mesh[counter6][1])
#                             counter6 = counter6 + 1
#                         DiseasePoint = []
#                         layer = layer + 1
#                         DiseasePoint.append(DiseaseName)
#                         DiseasePoint.append(IDGroup)
#                         DiseasePoint.append(layer)
#                         DAGs[counter].append(DiseasePoint)  ######################
#
#             counter2 = counter2 + 1
#         counter1 = counter1 + 1
#     counter = counter + 1
# print('DAGs', len(DAGs))
# storFile(DAGs, 'DAGs.csv')


# 构建model1
# 构建DV(disease value)，通过AllDisease构建的DiseaseAndMesh和DAGs，所以疾病顺序都一样，通过dags的layer构建DiseaseValue
# DiseaseValue = []
# counter = 0
# while counter < len(AllDisease):
#     if DAGs[counter][0][1] == 0:
#         DiseaseValuePair = []
#         DiseaseValuePair.append(AllDisease[counter])
#         DiseaseValuePair.append(0)
#         DiseaseValue.append(DiseaseValuePair)
#         counter = counter + 1
#         continue
#     counter1 = 0
#     DV = 0
#     while counter1 < len(DAGs[counter]):
#         DV = DV + math.pow(0.5, DAGs[counter][counter1][2])
#         counter1 = counter1 + 1
#     DiseaseValuePair = []
#     DiseaseValuePair.append(AllDisease[counter])
#     DiseaseValuePair.append(DV)
#     DiseaseValue.append(DiseaseValuePair)
#     counter = counter + 1
# print('len(DiseaseValue)', len(DiseaseValue))
# storFile(DiseaseValue, 'DiseaseValue.csv')


# 生成两个疾病DAGs相同部分的DV
# SameValue1 = []
# counter = 0
# while counter < len(AllDisease):
#     RowValue = []
#     if DiseaseValue[counter][1] == 0:           # 没有mesh id，整行都为0
#         counter1 = 0
#         while counter1 < len(AllDisease):
#             RowValue.append(0)
#             counter1 = counter1 + 1
#         SameValue1.append(RowValue)
#         counter = counter + 1
#         continue
#     counter1 = 0
#     while counter1 < len(AllDisease):#疾病counter和疾病counter1之间的共同节点
#         if DiseaseValue[counter1][1] == 0:  # 没有mesh id，此点为0
#             RowValue.append(0)
#             counter1 = counter1 + 1
#             continue
#         DiseaseAndDiseaseSimilarityValue = 0
#         counter2 = 0
#         while counter2 < len(DAGs[counter]):#疾病counter的所有DAGs的节点
#             counter3 = 0
#             while counter3 < len(DAGs[counter1]):#疾病counter1的所有DAGs的节点
#                 if DAGs[counter][counter2][0] == DAGs[counter1][counter3][0]:#找出共同节点
#                     DiseaseAndDiseaseSimilarityValue = DiseaseAndDiseaseSimilarityValue + math.pow(0.5, DAGs[counter][counter2][2]) + math.pow(0.5, DAGs[counter1][counter3][2]) #自己和自己的全部节点相同，对角线即DiseaseValue的两倍
#                 counter3 = counter3 + 1
#             counter2 = counter2 + 1
#         RowValue.append(DiseaseAndDiseaseSimilarityValue)
#         counter1 = counter1 + 1
#     SameValue1.append(RowValue)
#     counter = counter + 1
# print('SameValue1')
# storFile(SameValue1, 'Samevalue1.csv')


# 生成model1
# DiseaseSimilarityModel1 = []
# counter = 0
# while counter < len(AllDisease):
#     RowValue = []
#     if DiseaseValue[counter][1] == 0:           # 没有mesh id，整行都为0
#         counter1 = 0
#         while counter1 < len(AllDisease):
#             RowValue.append(0)
#             counter1 = counter1 + 1
#         DiseaseSimilarityModel1.append(RowValue)
#         counter = counter + 1
#         continue
#     counter1 = 0
#     while counter1 < len(AllDisease):
#         if DiseaseValue[counter1][1] == 0:  # 没有mesh id，此点为0
#             RowValue.append(0)
#             counter1 = counter1 + 1
#             continue
#         value = SameValue1[counter][counter1] / (DiseaseValue[counter][1] + DiseaseValue[counter1][1])
#         RowValue.append(value)
#         counter1 = counter1 + 1
#     DiseaseSimilarityModel1.append(RowValue)
#     counter = counter + 1
# print('DiseaseSimilarityModel1，行数', len(DiseaseSimilarityModel1))
# print('DiseaseSimilarityModel1[0]，列数', len(DiseaseSimilarityModel1[0]))
# storFile(DiseaseSimilarityModel1, 'DiseaseSimilarityModel1.csv')



# 构建model2
# 构建MeshAllDisease，mesh中的所有不相同疾病
# MeshAllDisease = []
# counter = 0
# while counter < len(mesh):
#     counter1 = 0
#     flag = 0
#     while counter1 < len(MeshAllDisease):
#         if mesh[counter][0] == MeshAllDisease[counter1]:
#             flag = 1
#             break
#         counter1 = counter1 + 1
#     if flag == 0:
#         MeshAllDisease.append(mesh[counter][0])
#     counter = counter + 1
# print('len(MeshAllDisease)', len(MeshAllDisease))
# storFile(MeshAllDisease, 'MeshAllDisease.csv')

# 构建MeshAllDiseaseAndMeshID
# MeshAllDiseaseAndMeshID = []
# counter1 = 0
# while counter1 < len(MeshAllDisease):
#     DiseaseAndMeshPair = []
#     MeshAllDiseaseAndMeshID.append(DiseaseAndMeshPair)
#     MeshAllDiseaseAndMeshID[counter1].append(MeshAllDisease[counter1])
#     counter2 = 0
#     flag = 0
#     while counter2 < len(mesh):#遍历整个mesh，寻找相同疾病的所有id
#         if (mesh[counter2][0] == MeshAllDiseaseAndMeshID[counter1][0]) & (flag == 1):#加入
#             MeshAllDiseaseAndMeshID[counter1][1].append(mesh[counter2][1])
#         if (mesh[counter2][0] == MeshAllDiseaseAndMeshID[counter1][0]) & (flag == 0):#新建mesh id 列表
#             MeshID = []
#             MeshID.append(mesh[counter2][1])
#             MeshAllDiseaseAndMeshID[counter1].append(MeshID)
#             flag = 1
#         counter2 = counter2 + 1
#     counter1 = counter1 + 1
# print('len(MeshAllDiseaseAndMeshID)', len(MeshAllDiseaseAndMeshID))
# storFile(MeshAllDiseaseAndMeshID, 'MeshAllDiseaseAndMeshID.csv')

# 构建MeshAllDiseaseDAGs的根节点
# MeshAllDiseaseDAGs = []
# counter1 = 0
# while counter1 < len(MeshAllDisease):
#     group = []
#     group.extend(MeshAllDiseaseAndMeshID[counter1])
#     group.append(0)
#     group1 = []
#     group1.append(group)
#     MeshAllDiseaseDAGs.append(group1)
#     counter1 = counter1 + 1
# print('MeshAllDiseaselen(DAGs)的叶子', len(MeshAllDiseaseDAGs))


# 构建MeshAllDiseaseDAGs
# counter = 0
# while counter < len(MeshAllDiseaseDAGs):
#     counter1 = 0
#     while counter1 < len(MeshAllDiseaseDAGs[counter]):  #################
#         counter2 = 0
#         while counter2 < len(MeshAllDiseaseDAGs[counter][counter1][1]):  ###################只对一个节点扩展只能生成的二层信息
#             layer = MeshAllDiseaseDAGs[counter][counter1][2]  #######################
#             # if len(DAGs[0][counter1][1][counter2]) <= 3:
#             #     break
#             if len(MeshAllDiseaseDAGs[counter][counter1][1][counter2]) > 3:  ####################
#                 NID = MeshAllDiseaseDAGs[counter][counter1][1][counter2]  #####################
#                 L = len(NID)
#                 NID = NID[0:L - 4]  # 把id减3
#                 counter3 = 0
#                 flag = 1  # 默认不在
#                 while counter3 < len(mesh):  # 判断nid是否在mesh中，如果在求出疾病名，如果不在，跳出循还
#                     if NID == mesh[counter3][1]:
#                         flag = 0  # 由counter3找对应的疾病名
#                         num = counter3
#                         DiseaseName = mesh[counter3][0]
#                         break
#                     counter3 = counter3 + 1
#
#                 DiseaseName = mesh[num][0]
#                 flag2 = 0  # 默认在dags不存在
#                 counter5 = 0
#                 while counter5 < len(MeshAllDiseaseDAGs[counter]):  # 找到对应疾病的名字后查找dags看是否已经出现，出现了就不加了
#                     if MeshAllDiseaseDAGs[counter][counter5][0] == DiseaseName:  #########################
#                         flag2 = 1  # dags中出现了
#                         break
#                     counter5 = counter5 + 1
#
#                 if flag == 0:
#                     if flag2 == 0:
#                         counter6 = 0    # 遍历mesh，寻找disease对应的id
#                         IDGroup = []
#                         while counter6 < len(mesh):
#                             if DiseaseName == mesh[counter6][0]:
#                                 IDGroup.append(mesh[counter6][1])
#                             counter6 = counter6 + 1
#                         DiseasePoint = []
#                         layer = layer + 1
#                         DiseasePoint.append(DiseaseName)
#                         DiseasePoint.append(IDGroup)
#                         DiseasePoint.append(layer)
#                         MeshAllDiseaseDAGs[counter].append(DiseasePoint)  ######################
#
#             counter2 = counter2 + 1
#         counter1 = counter1 + 1
#     counter = counter + 1
# print('len(MeshAllDiseaseDAGs)', len(MeshAllDiseaseDAGs))
# storFile(MeshAllDiseaseDAGs, 'MeshAllDiseaseDAGs.csv')

# 构建DiseaseFrequence，AllDisease在MeshAllDiseaseDAGs中出现的次数，可能为0次
# DiseaseFrequence = []
# counter = 0
# while counter < len(AllDisease):
#     num = 0
#     counter1 = 0
#     while counter1 < len(MeshAllDisease):#遍历所有疾病，疾病counter是否在疾病counter1中出现过
#         counter2 = 0
#         while counter2 < len(MeshAllDiseaseDAGs[counter1]):
#             if AllDisease[counter] == MeshAllDiseaseDAGs[counter1][counter2][0]:
#                 num = num + 1
#                 break
#             counter2 = counter2 + 1
#         counter1 = counter1 + 1
#     DiseaseFrequencePair = []
#     DiseaseFrequencePair.append(AllDisease[counter])
#     DiseaseFrequencePair.append(num)
#     DiseaseFrequence.append(DiseaseFrequencePair)
#     counter = counter + 1
# print('len(DiseaseFrequence)', len(DiseaseFrequence))
# storFile(DiseaseFrequence, 'DiseaseFrequence.csv')

# 计算每个疾病dags中的对数和，model2的分母DV2
# from math import e
# from math import log
# DiseaseValue2 = []
# counter = 0
# while counter < len(DiseaseFrequence):
#
#     if DiseaseFrequence[counter][1] == 0:   # 如果出现0次
#         DiseaseValuePair = []
#         DiseaseValuePair.append(AllDisease[counter])
#         DiseaseValuePair.append(0)
#         DiseaseValue2.append(DiseaseValuePair)  # [疾病,疾病对应的DAGs所有节点的DV2的和]
#         counter = counter + 1
#         continue
#     counter1 = 0
#     DV2 = 0
#     if DiseaseFrequence[counter][1] > 0:
#         while counter1 < len(DAGs[counter]):    # 找出DAGs[counter]疾病生成的dags所包含的疾病！！！！！！！！！！！！！！！看model1的找共同节点
#             DName = DAGs[counter][counter1][0]
#             counter2 = 0
#             while counter2 < len(DiseaseFrequence):
#                 if DiseaseFrequence[counter2][0] == DName:
#                     t = DiseaseFrequence[counter2][1] / len(MeshAllDisease)  # 有DAGs的生成
#                     DV2 = DV2 - log(t, 2)
#                 counter2 = counter2 + 1
#             counter1 = counter1 + 1
#         DiseaseValuePair = []
#         DiseaseValuePair.append(AllDisease[counter])
#         DiseaseValuePair.append(DV2)
#         DiseaseValue2.append(DiseaseValuePair)  # [疾病,疾病对应的DAGs所有节点的DV2的和]
#         counter = counter + 1
#
# print('len(DiseaseValue2)', len(DiseaseValue2))
# storFile(DiseaseValue2, 'DiseaseValue2.csv')

# DiseaseSimilarityModel2
# DiseaseSimilarityModel2 = []
# counter = 0
# while counter < len(AllDisease):
#     RowValue = []
#     if DiseaseFrequence[counter][1] == 0:
#         counter1= 0
#         while counter1 < len(AllDisease):
#             RowValue.append(0)
#             counter1 = counter1 + 1
#         DiseaseSimilarityModel2.append(RowValue)
#         counter = counter + 1
#         continue
#     counter1 = 0
#     while counter1 < len(AllDisease):#疾病counter和疾病counter1之间的共同节点
#         DiseaseAndDiseaseSimilarityValue = 0
#         if DiseaseFrequence[counter1][1] == 0:
#             RowValue.append(0)
#             counter1 = counter1 + 1
#             continue
#         counter2 = 0
#         while counter2 < len(DAGs[counter]):#疾病counter的所有DAGs的节点
#             counter3 = 0
#             while counter3 < len(DAGs[counter1]):#疾病counter1的所有DAGs的节点
#                 if DAGs[counter][counter2][0] == DAGs[counter1][counter3][0]:
#                     DName = DAGs[counter][counter2][0]  # 找共同疾病的出现的次数
#                     counter4 = 0
#                     while counter4 < len(DiseaseFrequence):
#                         if DName == DiseaseFrequence[counter4][0]:
#                             t1 = DiseaseFrequence[counter4][1] / len(MeshAllDisease)
#                             t2 = DiseaseFrequence[counter4][1] / len(MeshAllDisease)
#                             break
#                         counter4 = counter4 + 1
#                     DiseaseAndDiseaseSimilarityValue = DiseaseAndDiseaseSimilarityValue - log(t1, e) - log(t2, e)  # 通过循环，把共同节点的对数值加起来
#                 counter3 = counter3 + 1
#             counter2 = counter2 + 1
#         DiseaseAndDiseaseSimilarityValue = DiseaseAndDiseaseSimilarityValue / (DiseaseValue2[counter][1] + DiseaseValue2[counter1][1])
#         RowValue.append(DiseaseAndDiseaseSimilarityValue)
#         counter1 = counter1 + 1
#     DiseaseSimilarityModel2.append(RowValue)
#     counter = counter + 1
# print('len(DiseaseSimilarityModel2)', len(DiseaseSimilarityModel2))
# storFile(DiseaseSimilarityModel2, 'DiseaseSimilarityModel2.csv')


# disease的文本挖掘相似矩阵
lines = [line.strip().split() for line in open("disease相似性矩阵.txt")]
txtSimilarity = []
i = 0
for dis in lines:
    i = i + 1
    if i == 1:
        continue
    txtSimilarity.append(dis[1:])
print(len(txtSimilarity))
print(len(txtSimilarity[1]))


# 计算rd
counter1 = 0
sum1 = 0
while counter1 < (len(AllDisease)):
    counter2 = 0
    while counter2 < (len(AllDRUG)):
        sum1 = sum1 + pow((DiseaseAndDrugBinary[counter1][counter2]), 2)
        counter2 = counter2 + 1
    counter1 = counter1 + 1
print('sum1=', sum1)
Ak = sum1
Nd = len(AllDisease)
rdpie = 0.5
rd = rdpie * Nd / Ak
print('disease rd', rd)
# 生成DiseaseGaussian
DiseaseGaussian = []
counter1 = 0
while counter1 < len(AllDisease):#计算疾病counter1和counter2之间的similarity
    counter2 = 0
    DiseaseGaussianRow = []
    while counter2 < len(AllDisease):# 计算Ai*和Bj*
        AiMinusBj = 0
        sum2 = 0
        counter3 = 0
        AsimilarityB = 0
        while counter3 < len(AllDRUG):#疾病的每个属性分量
            sum2 = pow((DiseaseAndDrugBinary[counter1][counter3] - DiseaseAndDrugBinary[counter2][counter3]), 2)#计算平方
            AiMinusBj = AiMinusBj + sum2
            counter3 = counter3 + 1
        AsimilarityB = math.exp(- (AiMinusBj/rd))
        DiseaseGaussianRow.append(AsimilarityB)
        counter2 = counter2 + 1
    DiseaseGaussian.append(DiseaseGaussianRow)
    counter1 = counter1 + 1
print('len(DiseaseGaussian)', len(DiseaseGaussian))
# storFile(DiseaseGaussian, 'DiseaseGaussian.csv')



# 构建Drugaussian
from numpy import *
MDiseaseAndDrugBinary = np.array(DiseaseAndDrugBinary)    # 列表转为矩阵
DRUGAndDiseaseBinary = MDiseaseAndDrugBinary.T    # 转置DiseaseAndMiRNABinary
DRUGGaussian = []
counter1 = 0
sum1 = 0
while counter1 < (len(AllDRUG)):     # rna数量
    counter2 = 0
    while counter2 < (len(AllDisease)):     # disease数量
        sum1 = sum1 + pow((DRUGAndDiseaseBinary[counter1][counter2]), 2)
        counter2 = counter2 + 1
    counter1 = counter1 + 1
print('sum1=', sum1)
Ak = sum1
Nm = len(AllDRUG)
rdpie = 0.5
rd = rdpie * Nm / Ak
print('DRUG rd', rd)
# 生成DRUGGaussian
counter1 = 0
while counter1 < len(AllDRUG):   # 计算rna counter1和counter2之间的similarity
    counter2 = 0
    DRUGGaussianRow = []
    while counter2 < len(AllDRUG):   # 计算Ai*和Bj*
        AiMinusBj = 0
        sum2 = 0
        counter3 = 0
        AsimilarityB = 0
        while counter3 < len(AllDisease):   # rna的每个属性分量
            sum2 = pow((DRUGAndDiseaseBinary[counter1][counter3] - DRUGAndDiseaseBinary[counter2][counter3]), 2)#计算平方，有问题？？？？？
            AiMinusBj = AiMinusBj + sum2
            counter3 = counter3 + 1
        AsimilarityB = math.exp(- (AiMinusBj/rd))
        DRUGGaussianRow.append(AsimilarityB)
        counter2 = counter2 + 1
    DRUGGaussian.append(DRUGGaussianRow)
    counter1 = counter1 + 1
print('type(DRUGGaussian)', type(DRUGGaussian))
# storFile(DRUGGaussian, 'DRUGGaussian.csv')

# 把fingerprint降维
workbook = xlrd.open_workbook('newfingerprint.xlsx')
booksheet = workbook.sheet_by_index(0)
booksheet = workbook.sheet_by_name('Sheet1')

nrows = booksheet.nrows
row_list = []

for i in range(0, nrows):
         row=[]
         row_data = booksheet.row_values(i)
         ii = row_data
         for j in ii[0]:
             row.append(j)
         row_list.append(row)

y_test_test=[]
x_train_test=[]
x_test_test=[]
x_train_test_all=[]
x_all=[]
i=0
for row in row_list: #接口
    a = row
    x_train_test.append(a)
    i=i+1
x_train_test =np.array(x_train_test)
x_test_test =np.array(x_test_test)
(x_train, _), (x_test, y_test) = mnist.load_data()
x_train = x_train.astype('float32') / 255. - 0.5       # minmax_normalized
x_test = x_test.astype('float32') / 255. - 0.5         # minmax_normalized
x_train_test = x_train_test.reshape((x_train_test.shape[0], -1))
x_train = x_train_test
x_test = x_test_test

# in order to plot in a 2D figure
encoding_dim = 64

# this is our input placeholder
input_img = Input(shape=(920,))

# encoder layers
encoded = Dense(500, activation='relu')(input_img)
encoded = Dense(200, activation='relu')(encoded)
encoded = Dense(100, activation='relu')(encoded)
encoder_output = Dense(encoding_dim)(encoded)
print()
# decoder layers
decoded = Dense(100, activation='relu')(encoder_output)
decoded = Dense(200, activation='relu')(decoded)
decoded = Dense(500, activation='relu')(decoded)
decoded = Dense(920, activation='tanh')(decoded)

# construct the autoencoder model
autoencoder = Model(input=input_img, output=decoded)

# construct the encoder model for plotting
encoder = Model(input=input_img, output=encoder_output)

# compile autoencoder
autoencoder.compile(optimizer='adam', loss='mse')

# training
autoencoder.fit(x_train, x_train,
                epochs=10,
                batch_size=100,
                shuffle=True)

# plotting
encoded_imgs = encoder.predict(x_train)


# to xlsx
from openpyxl import Workbook

workbook = Workbook()
booksheet = workbook.active

drugSimilarity = []
for rows in encoded_imgs:
    booksheet.append(rows.tolist())
    drugSimilarity.append(rows.tolist())
print('drugSimilarity',len(drugSimilarity[0]))
# workbook.save("drugSimilarity.csv")






# 挑选正负样本
# 挑选正/负例
import random
counter1 = 0    # 在疾病中随机选择
counter2 = 0    # 在rna中随机选择
counterP = 0    # 正样本数量
counterN = 0    # 负样本数量
PositiveSample = []     # rna - disease 对
# 若正例为全部的RNA-Disease对
PositiveSample = LncDisease
print('PositiveSample)', len(PositiveSample))
# storFile(PositiveSample, 'PositiveSample.csv')


# 负样本为全部的disease-drug（313*593）中随机抽取，未在内LncDisease即为负样本
NegativeSample = []
counterN = 0
while counterN < len(PositiveSample):                         # 当正负样本任一小于10时执行循环，10用来测试，应与正样本数目相同，len(PositiveSample)！！！！！！！！！！！！！！！！！！！！
    counterD = random.randint(0, len(AllDisease)-1)
    counterR = random.randint(0, len(AllDRUG)-1)     # 随机选出一个疾病rna对
    DiseaseAndRnaPair = []
    DiseaseAndRnaPair.append(AllDRUG[counterR])
    DiseaseAndRnaPair.append(AllDisease[counterD])
    flag1 = 0
    counter = 0
    while counter < len(LncDisease):
        if DiseaseAndRnaPair == LncDisease[counter]:
            flag1 = 1
            break
        counter = counter + 1
    if flag1 == 1:
        continue
    flag2 = 0
    counter1 = 0
    while counter1 < len(NegativeSample):
        if DiseaseAndRnaPair == NegativeSample[counter1]:
            flag2 = 1
            break
        counter1 = counter1 + 1
    if flag2 == 1:
        continue
    if (flag1 == 0 & flag2 == 0):
        NegativePair = []
        NegativePair.append(AllDRUG[counterR])
        NegativePair.append(AllDisease[counterD])
        NegativeSample.append(NegativePair)
        counterN = counterN + 1
print('len(NegativeSample)', len(NegativeSample))
# storFile(NegativeSample, 'NegativeSample.csv')


# 由txtSimilarity，model1，model2，Gaussian生成最终的Similarity，有语义相似性，在model1/2矩阵中有值的就用model，没有的就用高斯，合成一个矩阵
DiseaseSimilarity = []
counter = 0
while counter < len(AllDisease):
    counter1 = 0
    Row = []
    while counter1 < len(AllDisease):
        v = float(DiseaseGaussian[counter][counter1])
        # v = (DiseaseSimilarityModel1[counter][counter1] + DiseaseSimilarityModel2[counter][counter1]) / 2
        if v > 0:
            Row.append(v)
        if v == 0:
            Row.append(txtSimilarity[counter][counter1])
        counter1 = counter1 + 1
    DiseaseSimilarity.append(Row)
    counter = counter + 1
print('len(DiseaseSimilarity)', len(DiseaseSimilarity))
print('len(DiseaseSimilarity[0)',len(DiseaseSimilarity[0]))
storFile(DiseaseSimilarity, 'DiseaseSimilarity.csv')



# 由Gaussian,fingerprint  drugSimilarity生成最终的Similarity，有语义相似性，在model1/2矩阵中有值的就用model，没有的就用高斯，合成一个矩阵
DRUGSimilarity = []
# counter = 0
# while counter < len(AllDRUG):
#     # counter1 = 0
#     Row = []
#     while counter1 < len(AllDRUG):
#         Row.append(DRUGGaussian[counter])
#         Row.extend(drugSimilarity[counter])
#     DRUGSimilarity.append(Row)
#     counter = counter + 1

DRUGSimilarity=np.hstack((DRUGGaussian, drugSimilarity))
print('len(DRUGSimilarity)', len(DRUGSimilarity))
print('len(DRUGSimilarity[0])', len(DRUGSimilarity[0]))
storFile(DRUGSimilarity, 'DRUGSimilarity.csv')


# 生成训练集 ！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！打乱顺序？？？
AllSample = PositiveSample.copy()
AllSample.extend(NegativeSample)        # drug-disease
# storFile(AllSample, 'AllSample.csv')

# SampleFeature
SampleFeature = []
counter = 0
while counter < len(AllSample):
    counter1 = 0
    while counter1 < len(AllDRUG):
        if AllSample[counter][0] == AllDRUG[counter1]:
            a = []
            counter3 = 0
            # 原本是ALLDrug
            while counter3 <len(DRUGSimilarity[0]):
                v = DRUGSimilarity[counter1][counter3]
                a.append(v)
                counter3 = counter3 + 1
            break
        counter1 = counter1 + 1
    counter2 = 0
    while counter2 < len(AllDisease):
        if AllSample[counter][1] == AllDisease[counter2]:
            b = []
            counter3 = 0
            # 原本是ALLDisease
            while counter3 < len(DiseaseSimilarity[0]):
                v = DiseaseSimilarity[counter2][counter3]
                b.append(v)
                counter3 = counter3 + 1
            break
        counter2 = counter2 + 1
    a.extend(b)
    SampleFeature.append(a)
    counter = counter + 1
counter1 = 0
storFile(SampleFeature, 'G64SampleFeature.csv')
print('SampleFeature',len(SampleFeature))
print('SampleFeature[1]',len(SampleFeature[1]))
# print('SampleFeature[1][1]',len(SampleFeature[1][1]))

# from openpyxl import Workbook
#
# workbook = Workbook()
# booksheet = workbook.active
#
# for rows in SampleFeature:
#     booksheet.append(rows)
#
# workbook.save("SampleFeature.xlsx")



# 打乱feature
# counter = 0
# R = []
# while counter < len(SampleFeature):
#     R.append(counter)
#     counter = counter + 1
# print(R)
# random.shuffle(R)
# print(R)
# RSampleFeature = []
# counter = 0
# print('len(SampleLabel)', len(SampleFeature))
# while counter < len(SampleFeature):
#     RSampleFeature.append(SampleFeature[R[counter]])
#     counter = counter + 1
# print('len(RSampleLabel)', len(RSampleFeature))
#
# SampleFeature = []
# SampleFeature = RSampleFeature


# # SampleLabel
# SampleLabel = []
# counter = 0
# while counter < len(PositiveSample):
#     SampleLabel.append(1)
#     counter = counter + 1
# print('len(SampleLabel)', len(SampleLabel))
# print('len(NegativeSample)', len(NegativeSample))
# counter1 = 0
# while counter1 < len(NegativeSample):
#     SampleLabel.append(0)
#     counter1 = counter1 + 1
# print('len(SampleLabel)', len(SampleLabel))
#

# row_list = []
# row_list = SampleFeature
#
#
#
# y_test_test=[]
# x_train_test=[]
# x_test_test=[]
# x_train_test_all=[]
# x_all=[]
# i=0
# for row in row_list: #接口
#     a = row
#     x_train_test.append(a)
#     i=i+1
# x_train_test =np.array(x_train_test)
# x_test_test =np.array(x_test_test)
# (x_train, _), (x_test, y_test) = mnist.load_data()
# x_train = x_train.astype('float32') / 255. - 0.5       # minmax_normalized
# x_test = x_test.astype('float32') / 255. - 0.5         # minmax_normalized
# x_train_test = x_train_test.reshape((x_train_test.shape[0], -1))
# x_train = x_train_test
# x_test = x_test_test
#
# # in order to plot in a 2D figure
# encoding_dim = 520
#
# # this is our input placeholder
# input_img = Input(shape=(1072,))
#
# # encoder layers
# encoded = Dense(500, activation='relu')(input_img)
# encoded = Dense(200, activation='relu')(encoded)
# encoded = Dense(100, activation='relu')(encoded)
# encoder_output = Dense(encoding_dim)(encoded)
# print()
# # decoder layers
# decoded = Dense(100, activation='relu')(encoder_output)
# decoded = Dense(200, activation='relu')(decoded)
# decoded = Dense(500, activation='relu')(decoded)
# decoded = Dense(1072, activation='tanh')(decoded)
#
# # construct the autoencoder model
# autoencoder = Model(input=input_img, output=decoded)
#
# # construct the encoder model for plotting
# encoder = Model(input=input_img, output=encoder_output)
#
# # compile autoencoder
# autoencoder.compile(optimizer='adam', loss='mse')
#
# # training
# autoencoder.fit(x_train, x_train,
#                 epochs=10,
#                 batch_size=500,
#                 shuffle=True)
#
# # plotting
# encoded_imgs = encoder.predict(x_train)
#
#
# # to xlsx
# from openpyxl import Workbook
#
# workbook = Workbook()
# booksheet = workbook.active
#
# SampleFeature = []
# for rows in encoded_imgs:
#     booksheet.append(rows.tolist())
#     SampleFeature.append(rows.tolist())
#
# # workbook.save("SampleFeature.csv")
# workbook.save("C520SampleFeature.xlsx")




# 打乱标签
# counter = 0
# R = []
# while counter < len(SampleFeature):
#     R.append(counter)
#     counter = counter + 1
# print(R)
# random.shuffle(R)
# print(R)
# RSampleLabel = []
# RSampleFeature = []
# counter = 0
# print('len(SampleLabel)', len(SampleLabel))
# while counter < len(SampleLabel):
#     RSampleLabel.append(SampleLabel[R[counter]])
#     # RSampleFeature.append(SampleFeature[counter])
#     counter = counter + 1
# print(RSampleLabel)

# SampleLabel = []
# SampleLabel = RSampleLabel

# storFile(RSampleLabel, 'RSampleLabel.csv')        # 保存csv会报错，改变数据类型



# SampleFeature = np.array(SampleFeature)
# SampleLabel = np.array(SampleLabel)
# print(len(SampleFeature))
# print(len(SampleLabel))
